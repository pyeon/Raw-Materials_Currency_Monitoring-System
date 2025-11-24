"""
엑셀 리포트 생성 모듈
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from typing import Dict
from config import get_enabled_assets, REPORT_DIR, EXCEL_CONFIG

class ExcelReporter:
    """엑셀 리포트 생성 클래스"""
    
    def __init__(self, processed_data: Dict):
        self.data = processed_data
        self.assets = get_enabled_assets()
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        
    def generate_report(self) -> str:
        """전체 리포트 생성"""
        # 1. 종합 요약 시트
        self._create_summary_sheet()
        
        # 2. 일자별 상세 시트
        self._create_daily_detail_sheet()
        
        # 3. 주간 추이 시트
        self._create_weekly_trend_sheet()
        
        # 4. 월간 추이 시트
        self._create_monthly_trend_sheet()
        
        # 5. 기술적 지표 시트
        self._create_technical_indicators_sheet()
        
        # 6. 상관관계 시트
        self._create_correlation_sheet()
        
        # 파일 저장
        os.makedirs(REPORT_DIR, exist_ok=True)
        filename = f"commodity_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(REPORT_DIR, filename)
        
        self.workbook.save(filepath)
        print(f"✅ 엑셀 리포트 생성: {filepath}")
        
        return filepath
    
    def _create_summary_sheet(self):
        """종합 요약 시트"""
        ws = self.workbook.create_sheet("📊 종합요약", 0)
        
        # 헤더
        headers = ['구분', '자산', '현재가', '전일비', '변동률(%)', 
                   '주간변동(%)', '월간변동(%)', '52주최고', '52주최저', '추세']
        ws.append(headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        
        # 모든 카테고리 처리
        category_names = {
            'commodities': '원자재',
            'currencies': '통화',
            'cryptocurrencies': '암호화폐'
        }
        
        for category, assets in self.assets.items():
            for code, info in assets.items():
                if code not in self.data or 'error' in self.data[code]:
                    continue
                
                d = self.data[code]
                
                # 주간/월간 변동률 계산
                weekly_change = self._calculate_period_change(d, 'weekly')
                monthly_change = self._calculate_period_change(d, 'monthly')
                
                # 추세 판단
                trend = self._determine_trend(d)
                
                row_data = [
                    category_names.get(category, category),
                    info['name'],
                    d['current_price'],
                    d['daily_change'],
                    d['daily_change_pct'],
                    weekly_change,
                    monthly_change,
                    d.get('52w_high', '-'),
                    d.get('52w_low', '-'),
                    trend
                ]
                
                ws.append(row_data)
                
                # 변동률에 따른 색상
                self._apply_change_color(ws, row_num, 5, d['daily_change_pct'])
                row_num += 1
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
    
    def _create_daily_detail_sheet(self):
        """일자별 상세 시트"""
        ws = self.workbook.create_sheet("📅 일자별상세")
        
        # 헤더 생성
        header_row = ['날짜']
        for category, assets in self.assets.items():
            for code, info in assets.items():
                if code in self.data and 'error' not in self.data[code]:
                    header_row.append(info['name'])
        
        ws.append(header_row)
        
        # 최근 7일 데이터 추출
        all_dates = set()
        for code, d in self.data.items():
            if code != 'correlations' and 'last_7days' in d:
                all_dates.update(d['last_7days'].keys())
        
        sorted_dates = sorted(all_dates, reverse=True)[:7]
        
        for date in sorted_dates:
            row_data = [date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)]
            
            for category, assets in self.assets.items():
                for code in assets.keys():
                    if code in self.data and code != 'correlations':
                        d = self.data[code]
                        if 'last_7days' in d and date in d['last_7days']:
                            row_data.append(d['last_7days'][date])
                        else:
                            row_data.append('-')
            
            ws.append(row_data)
        
        # 스타일링
        self._style_header_row(ws, 1)
    
    def _create_weekly_trend_sheet(self):
        """주간 추이 시트"""
        ws = self.workbook.create_sheet("📈 주간추이")
        
        headers = ['자산', '당주평균', '전주평균', '전전주평균', '최근4주평균', 
                   '전주대비(%)', '전전주대비(%)']
        ws.append(headers)
        self._style_header_row(ws, 1)
        
        for category, assets in self.assets.items():
            for code, info in assets.items():
                if code not in self.data or 'error' in self.data[code]:
                    continue
                
                d = self.data[code]
                weekly = d.get('weekly', {})
                
                current = weekly.get('current_period_avg', '-')
                last = weekly.get('last_period_avg', '-')
                last_2 = weekly.get('last_2_period_avg', '-')
                last_4 = weekly.get('last_3month_avg', '-')
                
                # 변동률 계산
                last_change = ((current - last) / last * 100) if isinstance(current, (int, float)) and isinstance(last, (int, float)) else '-'
                last_2_change = ((current - last_2) / last_2 * 100) if isinstance(current, (int, float)) and isinstance(last_2, (int, float)) else '-'
                
                row_data = [
                    info['name'],
                    current,
                    last,
                    last_2,
                    last_4,
                    last_change,
                    last_2_change
                ]
                
                ws.append(row_data)
        
        # 컬럼 너비
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_monthly_trend_sheet(self):
        """월간 추이 시트"""
        ws = self.workbook.create_sheet("📊 월간추이")
        
        headers = ['자산', '당월평균', '전월평균', '전전월평균', 
                   '최근3개월', '최근6개월', '최근12개월', '전월대비(%)']
        ws.append(headers)
        self._style_header_row(ws, 1)
        
        for category, assets in self.assets.items():
            for code, info in assets.items():
                if code not in self.data or 'error' in self.data[code]:
                    continue
                
                d = self.data[code]
                monthly = d.get('monthly', {})
                
                current = monthly.get('current_period_avg', '-')
                last = monthly.get('last_period_avg', '-')
                last_2 = monthly.get('last_2_period_avg', '-')
                last_3 = monthly.get('last_3month_avg', '-')
                last_6 = monthly.get('last_6month_avg', '-')
                last_12 = monthly.get('last_12month_avg', '-')
                
                # 전월 대비 변동률
                last_change = ((current - last) / last * 100) if isinstance(current, (int, float)) and isinstance(last, (int, float)) else '-'
                
                row_data = [
                    info['name'],
                    current,
                    last,
                    last_2,
                    last_3,
                    last_6,
                    last_12,
                    last_change
                ]
                
                ws.append(row_data)
        
        # 컬럼 너비
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_technical_indicators_sheet(self):
        """기술적 지표 시트"""
        ws = self.workbook.create_sheet("🔧 기술지표")
        
        headers = ['자산', 'MA5', 'MA20', 'MA60', 'MA120', 
                   'MA5괴리(%)', 'MA20괴리(%)', '크로스신호', '배열상태']
        ws.append(headers)
        self._style_header_row(ws, 1)
        
        for category, assets in self.assets.items():
            for code, info in assets.items():
                if code not in self.data or 'error' in self.data[code]:
                    continue
                
                d = self.data[code]
                ma = d.get('moving_averages', {})
                signals = d.get('cross_signals', {})
                
                # 이동평균값
                ma5 = ma.get('MA5', {}).get('value', '-')
                ma20 = ma.get('MA20', {}).get('value', '-')
                ma60 = ma.get('MA60', {}).get('value', '-')
                ma120 = ma.get('MA120', {}).get('value', '-')
                
                # 괴리율
                ma5_div = ma.get('MA5', {}).get('divergence', '-')
                ma20_div = ma.get('MA20', {}).get('divergence', '-')
                
                # 크로스 신호
                cross_signal = ''
                if signals.get('golden_cross_5_20'):
                    cross_signal = '골든크로스'
                elif signals.get('dead_cross_5_20'):
                    cross_signal = '데드크로스'
                else:
                    cross_signal = '-'
                
                # 배열 상태
                alignment = ''
                if signals.get('bullish_alignment'):
                    alignment = '정배열'
                elif signals.get('bearish_alignment'):
                    alignment = '역배열'
                else:
                    alignment = '-'
                
                row_data = [
                    info['name'],
                    ma5,
                    ma20,
                    ma60,
                    ma120,
                    ma5_div,
                    ma20_div,
                    cross_signal,
                    alignment
                ]
                
                ws.append(row_data)
        
        # 컬럼 너비
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 14
    
    def _create_correlation_sheet(self):
        """상관관계 시트"""
        ws = self.workbook.create_sheet("🔗 상관관계")
        
        if 'correlations' not in self.data:
            ws.append(['상관관계 데이터 없음'])
            return
        
        correlations = self.data['correlations']
        
        # 헤더
        ws.append(['자산 쌍', '상관계수', '관계 강도'])
        self._style_header_row(ws, 1)
        
        # 데이터 추가
        for pair, corr in sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True):
            # 자산명 변환
            codes = pair.split('_')
            names = []
            for code in codes:
                for category, assets in self.assets.items():
                    if code in assets:
                        names.append(assets[code]['name'])
                        break
                else:
                    names.append(code)
            
            pair_name = ' vs '.join(names)
            
            # 관계 강도 판단
            if abs(corr) > 0.7:
                strength = '강함'
            elif abs(corr) > 0.4:
                strength = '중간'
            else:
                strength = '약함'
            
            if corr > 0:
                strength += ' (정상관)'
            else:
                strength += ' (역상관)'
            
            row_data = [pair_name, corr, strength]
            ws.append(row_data)
        
        # 컬럼 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
    
    def _calculate_period_change(self, data: Dict, period: str) -> float:
        """기간별 변동률 계산"""
        period_data = data.get(period, {})
        current = data.get('current_price')
        last = period_data.get('last_period_avg')
        
        if current and last:
            return ((current - last) / last) * 100
        return 0.0
    
    def _determine_trend(self, data: Dict) -> str:
        """추세 판단"""
        ma = data.get('moving_averages', {})
        signals = data.get('cross_signals', {})
        
        if signals.get('bullish_alignment'):
            return '강세 (정배열)'
        elif signals.get('bearish_alignment'):
            return '약세 (역배열)'
        elif 'MA5' in ma and 'MA20' in ma:
            if ma['MA5']['position'] == 'above' and ma['MA20']['position'] == 'above':
                return '상승 추세'
            elif ma['MA5']['position'] == 'below' and ma['MA20']['position'] == 'below':
                return '하락 추세'
        
        return '보합'
    
    def _apply_change_color(self, ws, row: int, col: int, value: float):
        """변동률에 따른 색상 적용"""
        cell = ws.cell(row=row, column=col)
        
        if value > 2:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            cell.font = Font(color="FF0000", bold=True)
        elif value > 1:
            cell.fill = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")
            cell.font = Font(color="FF6600")
        elif value < -2:
            cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            cell.font = Font(color="0000FF", bold=True)
        elif value < -1:
            cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            cell.font = Font(color="0066FF")
    
    def _style_header_row(self, ws, row: int):
        """헤더 행 스타일 적용"""
        for cell in ws[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
